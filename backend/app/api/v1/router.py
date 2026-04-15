import uuid
import os
from io import BytesIO
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from pydantic import ValidationError

from app.db.session import get_db
from app.db import crud
from app.api.v1.schemas import (
    ConfigRequest,
    ConfigResponse,
    StatusResponse,
    HistoryResponse,
    HistoryItem,
    ConfigResult,
    KnowledgeBaseDocumentCreate,
    KnowledgeBaseDocumentUpdate,
    KnowledgeBaseDocumentResponse,
    KnowledgeBaseDocumentListResponse,
    KnowledgeBaseSearchRequest,
    KnowledgeBaseSearchResponse,
    FileUploadResponse,
    ExcelExportRequest
)
from app.config import settings

router = APIRouter()


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True, color="76B900")
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def _style_header_row(row) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell.alignment = WRAP_ALIGNMENT
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 60))
        worksheet.column_dimensions[column_letter].width = max(14, max_length + 2)


def build_excel_workbook(payload: ExcelExportRequest) -> BytesIO:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "配置概览"

    input_config = payload.input_config or {}
    result = payload.result or {}
    nic_recommendation = result.get("nic_recommendation", {})
    network_design = result.get("network_design", {})
    validation_report = result.get("validation_report", {})

    summary_sheet.append(["模块", "内容"])
    _style_header_row(summary_sheet[1])
    summary_rows = [
        ("服务器数量", input_config.get("server_count", "")),
        ("应用场景", input_config.get("use_case", "")),
        ("组网路线", input_config.get("fabric_type", "")),
        ("GPU 配置", " / ".join(
            f"{gpu.get('model', '')} x{gpu.get('count', '')} ({gpu.get('interconnect', '')})"
            for gpu in input_config.get("gpu_configs", [])
        )),
        ("推荐网卡型号", nic_recommendation.get("full_model", "")),
        ("每台网卡数量", nic_recommendation.get("count_per_server", "")),
        ("技术路线交换机", network_design.get("switch_type", "")),
        ("GPU 输入摘要", network_design.get("gpu_context_summary", "")),
        ("带宽级别判断", network_design.get("bandwidth_tier", "")),
        ("带宽分析", network_design.get("bandwidth_analysis_summary", "")),
        ("单机对外网络带宽", network_design.get("single_server_network_bandwidth", "")),
        ("性能预估", " / ".join(
            f"{key}: {value}" for key, value in (network_design.get("performance_estimate", {}) or {}).items()
        )),
        ("高可用设计", network_design.get("high_availability_design", "")),
        ("扩展性建议", network_design.get("scalability_suggestions", "")),
        ("校验评分", validation_report.get("validation_score", "")),
    ]
    for row in summary_rows:
        summary_sheet.append(row)

    rationale_sheet = workbook.create_sheet("配置依据")
    rationale_sheet.append(["类型", "说明"])
    _style_header_row(rationale_sheet[1])
    for item in network_design.get("nic_sizing_rationale", []) or []:
        rationale_sheet.append(["网卡数量依据", item])
    for item in network_design.get("cabling_guidance", []) or []:
        rationale_sheet.append(["链路规划提示", item])
    for item in nic_recommendation.get("planning_notes", []) or []:
        rationale_sheet.append(["推荐说明", item])
    for item in validation_report.get("optimization_suggestions", []) or []:
        rationale_sheet.append(["优化建议", item])
    rationale_text = validation_report.get("configuration_rationale")
    if rationale_text:
        rationale_sheet.append(["校验依据", rationale_text])

    bom_sheet = workbook.create_sheet("BOM清单")
    bom_sheet.append(["类型", "名称", "完整型号", "数量", "端口数", "端口类型", "速率", "距离", "线缆类型", "长度", "接口类型", "链接"])
    _style_header_row(bom_sheet[1])
    for item in result.get("bom_list", []) or []:
        bom_sheet.append([
            item.get("type", ""),
            item.get("name", ""),
            item.get("full_model", ""),
            item.get("quantity", ""),
            item.get("port_count", ""),
            item.get("port_type", ""),
            item.get("speed", ""),
            item.get("distance", ""),
            item.get("cable_type", ""),
            item.get("length", ""),
            item.get("interface_type", ""),
            item.get("link", ""),
        ])

    validation_sheet = workbook.create_sheet("校验结果")
    validation_sheet.append(["严重级别", "问题", "建议"])
    _style_header_row(validation_sheet[1])
    for issue in validation_report.get("issues", []) or []:
        validation_sheet.append([
            issue.get("severity", ""),
            issue.get("message", ""),
            issue.get("suggestion", ""),
        ])

    for sheet in workbook.worksheets:
        _autosize_columns(sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def process_configuration_task(task_id: str, db: Session):
    try:
        from app.chains.nic_recommendation_chain import get_nic_recommendation
        from app.chains.network_design_chain import get_network_design
        from app.chains.bom_generation_chain import get_bom_list
        from app.chains.validation_chain import get_validation_report
        
        config_history = crud.get_config_history_by_task_id(db, task_id)
        if not config_history:
            return
        
        input_config = config_history.input_config
        
        try:
            nic_recommendation = get_nic_recommendation(input_config)
            network_design = get_network_design(input_config, nic_recommendation)
            bom_list = get_bom_list(network_design, nic_recommendation, input_config)
            validation_report = get_validation_report(
                input_config,
                nic_recommendation,
                network_design,
                bom_list
            )
            
            crud.update_config_history_result(
                db,
                task_id=task_id,
                nic_recommendation=nic_recommendation.dict() if hasattr(nic_recommendation, 'dict') else nic_recommendation,
                network_design=network_design.dict() if hasattr(network_design, 'dict') else network_design,
                bom_list=[item.dict() if hasattr(item, 'dict') else item for item in bom_list],
                validation_report=validation_report.dict() if hasattr(validation_report, 'dict') else validation_report,
                status="completed"
            )
        except Exception as e:
            crud.update_config_history_status(
                db,
                task_id=task_id,
                status="failed",
                error_message=str(e)
            )
    except Exception as e:
        crud.update_config_history_status(
            db,
            task_id=task_id,
            status="failed",
            error_message=f"Task processing error: {str(e)}"
        )


@router.post("/configure", response_model=ConfigResponse)
async def submit_configuration(
    request: ConfigRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    task_id = str(uuid.uuid4())
    
    crud.create_config_history(
        db,
        task_id=task_id,
        input_config=request.dict(),
        session_id=request.session_id
    )
    
    from app.db.session import SessionLocal
    background_tasks.add_task(process_configuration_task, task_id, SessionLocal())
    
    return ConfigResponse(
        task_id=task_id,
        status="processing",
        message="Configuration task has been submitted successfully"
    )


@router.get("/status/{task_id}", response_model=StatusResponse)
async def get_configuration_status(
    task_id: str,
    db: Session = Depends(get_db)
):
    config_history = crud.get_config_history_by_task_id(db, task_id)
    
    if not config_history:
        raise HTTPException(status_code=404, detail="Task not found")
    
    result = None
    if config_history.status == "completed":
        try:
            result = ConfigResult(
                nic_recommendation=config_history.nic_recommendation,
                network_design=config_history.network_design,
                bom_list=config_history.bom_list,
                validation_report=config_history.validation_report
            )
        except ValidationError:
            result = None
    
    error = config_history.error_message if config_history.status == "failed" else None
    
    return StatusResponse(
        status=config_history.status,
        result=result,
        error=error
    )


@router.get("/history", response_model=HistoryResponse)
async def get_configuration_history(
    session_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    if session_id:
        history_items = crud.get_config_history_by_session_id(
            db, session_id, skip=skip, limit=limit
        )
        total_count = crud.count_config_history_by_session_id(db, session_id)
    else:
        history_items = crud.get_all_config_history(db, skip=skip, limit=limit)
        total_count = len(history_items)
    
    history_list = [
        HistoryItem(
            id=item.id,
            created_at=item.created_at,
            input_config=item.input_config,
            session_id=item.session_id
        )
        for item in history_items
    ]
    
    return HistoryResponse(
        history=history_list,
        total_count=total_count
    )


@router.post("/export/excel")
async def export_configuration_excel(payload: ExcelExportRequest):
    try:
        file_buffer = build_excel_workbook(payload)
        filename_prefix = payload.filename_prefix or "nvidia-config"
        filename = f"{filename_prefix}.xlsx"
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
        return StreamingResponse(
            file_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to export Excel: {exc}")


# ==================== 知识库管理 API ====================


@router.post("/knowledge-base/documents", response_model=KnowledgeBaseDocumentResponse)
async def create_knowledge_base_document(
    request: KnowledgeBaseDocumentCreate,
    db: Session = Depends(get_db)
):
    try:
        db_document = crud.create_knowledge_base_entry(
            db,
            source_url=request.source_url,
            title=request.title,
            category=request.category,
            content=request.content,
            vector_id=None
        )
        
        return KnowledgeBaseDocumentResponse(
            id=db_document.id,
            title=db_document.title,
            content=db_document.content,
            category=db_document.category,
            source=db_document.category,
            source_url=db_document.source_url,
            product_models=request.product_models,
            keywords=request.keywords,
            last_updated=db_document.last_updated,
            version=request.version,
            vector_id=db_document.vector_id
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create document: {str(e)}")


@router.get("/knowledge-base/documents", response_model=KnowledgeBaseDocumentListResponse)
async def list_knowledge_base_documents(
    category: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db)
):
    try:
        skip = (page - 1) * page_size
        documents = crud.get_knowledge_base_entries(
            db,
            category=category,
            skip=skip,
            limit=page_size
        )
        
        total_count = len(documents)
        
        document_responses = [
            KnowledgeBaseDocumentResponse(
                id=doc.id,
                title=doc.title or "",
                content=doc.content or "",
                category=doc.category or "",
                source=doc.category or "",
                source_url=doc.source_url,
                product_models=None,
                keywords=None,
                last_updated=doc.last_updated,
                version=None,
                vector_id=doc.vector_id
            )
            for doc in documents
        ]
        
        return KnowledgeBaseDocumentListResponse(
            documents=document_responses,
            total_count=total_count,
            page=page,
            page_size=page_size
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list documents: {str(e)}")


@router.get("/knowledge-base/documents/{document_id}", response_model=KnowledgeBaseDocumentResponse)
async def get_knowledge_base_document(
    document_id: int,
    db: Session = Depends(get_db)
):
    document = crud.get_knowledge_base_by_vector_id(db, str(document_id))
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    return KnowledgeBaseDocumentResponse(
        id=document.id,
        title=document.title or "",
        content=document.content or "",
        category=document.category or "",
        source=document.category or "",
        source_url=document.source_url,
        product_models=None,
        keywords=None,
        last_updated=document.last_updated,
        version=None,
        vector_id=document.vector_id
    )


@router.put("/knowledge-base/documents/{document_id}", response_model=KnowledgeBaseDocumentResponse)
async def update_knowledge_base_document(
    document_id: int,
    request: KnowledgeBaseDocumentUpdate,
    db: Session = Depends(get_db)
):
    document = crud.get_knowledge_base_by_vector_id(db, str(document_id))
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    try:
        if request.title:
            document.title = request.title
        if request.content:
            document.content = request.content
        if request.category:
            document.category = request.category
        if request.source_url:
            document.source_url = request.source_url
        
        db.commit()
        db.refresh(document)
        
        return KnowledgeBaseDocumentResponse(
            id=document.id,
            title=document.title or "",
            content=document.content or "",
            category=document.category or "",
            source=document.category or "",
            source_url=document.source_url,
            product_models=request.product_models,
            keywords=request.keywords,
            last_updated=document.last_updated,
            version=request.version,
            vector_id=document.vector_id
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update document: {str(e)}")


@router.delete("/knowledge-base/documents/{document_id}")
async def delete_knowledge_base_document(
    document_id: int,
    db: Session = Depends(get_db)
):
    document = crud.get_knowledge_base_by_vector_id(db, str(document_id))
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    try:
        db.delete(document)
        db.commit()
        return {"status": "ok", "message": "Document deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete document: {str(e)}")


@router.post("/knowledge-base/search", response_model=KnowledgeBaseSearchResponse)
async def search_knowledge_base(
    request: KnowledgeBaseSearchRequest,
    db: Session = Depends(get_db)
):
    try:
        documents = crud.get_knowledge_base_entries(
            db,
            category=request.category,
            skip=0,
            limit=request.top_k
        )
        
        results = [
            KnowledgeBaseDocumentResponse(
                id=doc.id,
                title=doc.title or "",
                content=doc.content or "",
                category=doc.category or "",
                source=doc.category or "",
                source_url=doc.source_url,
                product_models=None,
                keywords=None,
                last_updated=doc.last_updated,
                version=None,
                vector_id=doc.vector_id
            )
            for doc in documents
        ]
        
        return KnowledgeBaseSearchResponse(
            query=request.query,
            results=results,
            scores=None
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Search failed: {str(e)}")


@router.post("/knowledge-base/upload", response_model=FileUploadResponse)
async def upload_file(
    file: UploadFile = File(...),
    category: Optional[str] = None,
    db: Session = Depends(get_db)
):
    try:
        file_id = str(uuid.uuid4())
        file_extension = os.path.splitext(file.filename)[1].lower()
        
        upload_dir = os.path.join(os.getcwd(), "uploads")
        os.makedirs(upload_dir, exist_ok=True)
        
        file_path = os.path.join(upload_dir, f"{file_id}{file_extension}")
        
        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)
        
        db_document = crud.create_knowledge_base_entry(
            db,
            source_url=file_path,
            title=file.filename,
            category=category or "general",
            content=f"File uploaded: {file.filename}",
            vector_id=None
        )
        
        return FileUploadResponse(
            file_id=file_id,
            filename=file.filename,
            file_size=len(content),
            content_type=file.content_type or "application/octet-stream",
            upload_status="success",
            message=f"File uploaded successfully. Document ID: {db_document.id}"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"File upload failed: {str(e)}")
